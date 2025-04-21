from openpyxl import load_workbook, Workbook
import os

# Function to read data (excluding header) from a given Excel file
def reader(file):
    abs_file = os.path.join(path, file)
    wb_sheet = load_workbook(abs_file).active
    rows = []
    
    # Skip the header row (assumed at row 1)
    for row in wb_sheet.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    
    return rows

# Custom headers for the unified Excel file
headers = ['Nume', 'Prenume', 'Titlu', 'Editura', 'Cota', 'Pret', 'An']

# Ask user for name of final workbook
workbook_name = input('Unified Workbook name (with .xlsx): ')
book = Workbook()
sheet = book.active

# Write headers to the new workbook
sheet.append(headers)

# Ask user for path containing Excel files to merge
path = input('Path to folder containing Excel files: ')

# List all files in the directory
files = os.listdir(path)

# Loop through each file and add its rows to the unified workbook
for file in files:
    if file.endswith('.xlsx'):
        print(f"Processing: {file}")
        rows = reader(file)
        for row in rows:
            sheet.append(row)

# Save the final merged Excel workbook
book.save(filename=workbook_name)
print(f"✅ All files merged successfully into {workbook_name}")
